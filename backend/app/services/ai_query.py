from __future__ import annotations

import csv
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

import httpx

from app.config import settings
from app.models import DataSource
from app.services.diagnostics import (
    ADVISORY_SYSTEM,
    DIAGNOSTIC_SYSTEM,
    PARTIAL_SYSTEM,
    build_evidence_prompt,
    build_partial_prompt,
    parse_advisory_json,
)
from app.services.response_planner import (
    NARRATIVE_SYSTEM,
    build_narrative_prompt,
    expand_question_with_context,
    question_prompt_block,
    sanitize_narrative,
)
from app.services.schema_context import (
    build_schema_prompt,
    build_workspace_schema_prompt,
    heuristic_plan,
    pick_source_for_question,
)
from app.services.schema_registry import parse_connection_config
from app.services.sql_sandbox import ensure_limit, validate_readonly_sql

#: Mode reported when the offline planner fell back to a plain row dump. The
#: rows are real, but they do not answer the question, and the answer says so.
UNTARGETED = "heuristic-untargeted"

DATE_RULES = """
Working with dates:
- Column hints in square brackets give the REAL range of values in the data.
- Never invent a year. If the question names months or a season without a year
  (e.g. "between March and May"), use the year(s) shown in the date range.
- If the data spans several years and the question gives no year, cover them all
  rather than picking one (e.g. filter on the month, not on a single year).
- Dates are stored as text in ISO form, so string comparison works:
  WHERE d >= '2026-03-01' AND d < '2026-06-01'.
"""

AGGREGATION_RULES = """
Answering "how much / how many / total / average":
- Return the aggregate itself, not the underlying rows.
- Give the aggregate a readable alias, e.g. SUM(revenue) AS total_revenue.
"""

SYSTEM_PROMPT = (
    """You are a SQL expert for a business intelligence tool.
Given a database schema and a user question, return ONLY a single read-only SQL query.
Rules:
- SELECT or WITH only. No writes, DDL, or multiple statements.
- Use only tables/columns from the schema.
- Prefer LIMIT 50 unless the user asks for a specific limit.
- For MySQL / SQLite, use backticks for identifiers when needed.
- Do not wrap the SQL in markdown fences.
- Never answer questions about yourself, which model you are, or how the product
  works with SQL. Those are not data questions — if asked, return exactly:
  SELECT NULL AS unsupported WHERE 0;
- When a PREVIOUS QUESTION is provided, treat short follow-ups
  (e.g. "what about the least?", "and by region?") as continuing that analysis:
  keep the same measure, dimensions, and time filters unless the follow-up
  clearly changes them. Flip ranking words (highest↔lowest, most↔least, top↔bottom).
"""
    + DATE_RULES
    + AGGREGATION_RULES
)

WORKSPACE_SYSTEM_PROMPT = (
    """You are a SQL expert for a business intelligence workspace with multiple datasets.
The user asks questions against ALL ingested data. You must:
1. Choose exactly ONE SOURCE_ID from the catalog that best answers the question.
2. Write a single read-only SQL query for that source only (no cross-source joins).

Respond with EXACTLY this format (no markdown fences):
SOURCE_ID: <id>
SQL:
<single SELECT or WITH query>

Rules:
- SELECT or WITH only. No writes, DDL, or multiple statements.
- Use only tables/columns from the chosen source.
- Prefer LIMIT 50 unless the user asks for a specific limit.
- Never answer questions about yourself, which model you are, or how the product
  works with SQL. Those are not data questions — if asked, pick any SOURCE_ID and
  return: SELECT NULL AS unsupported WHERE 0;
- When a PREVIOUS QUESTION is provided, treat short follow-ups as continuing that
  analysis (same measure, dimensions, and time filters unless clearly changed).
  Flip ranking words (highest↔lowest, most↔least, top↔bottom).
"""
    + DATE_RULES
    + AGGREGATION_RULES
)


async def generate_sql(
    source: DataSource,
    question: str,
    *,
    api_key: str | None = None,
    model: str | None = None,
    base_url: str | None = None,
    previous_question: str | None = None,
) -> tuple[str, str]:
    """Returns (sql, mode) where mode is 'openai' or 'heuristic'."""
    key = api_key if api_key is not None else settings.openai_api_key
    mdl = model or settings.openai_model
    url = base_url or settings.openai_base_url
    resolved = expand_question_with_context(question, previous_question)

    if key:
        sql = await _openai_sql(
            source,
            question,
            api_key=key,
            model=mdl,
            base_url=url,
            previous_question=previous_question,
        )
        return sql, "openai"

    plan = heuristic_plan(source, resolved)
    if not plan["sql"]:
        raise ValueError(
            "No OpenAI API key configured and could not build a heuristic query. "
            "Set it in Settings or OPENAI_API_KEY in .env"
        )
    return plan["sql"], ("heuristic" if plan["targeted"] else UNTARGETED)


async def generate_workspace_sql(
    sources: list[DataSource],
    question: str,
    *,
    api_key: str | None = None,
    model: str | None = None,
    base_url: str | None = None,
    previous_question: str | None = None,
) -> tuple[DataSource, str, str]:
    """Pick a source + SQL for a workspace question. Returns (source, sql, mode)."""
    if not sources:
        raise ValueError("No data sources ingested yet. Add a dataset first.")

    key = api_key if api_key is not None else settings.openai_api_key
    mdl = model or settings.openai_model
    url = base_url or settings.openai_base_url
    resolved = expand_question_with_context(question, previous_question)

    if key and len(sources) > 1:
        source_id, sql = await _openai_workspace_sql(
            sources,
            question,
            api_key=key,
            model=mdl,
            base_url=url,
            previous_question=previous_question,
        )
        by_id = {s.id: s for s in sources}
        source = by_id.get(source_id) or pick_source_for_question(sources, resolved) or sources[0]
        return source, sql, "openai"

    if key:
        source = pick_source_for_question(sources, resolved) or sources[0]
        sql = await _openai_sql(
            source,
            question,
            api_key=key,
            model=mdl,
            base_url=url,
            previous_question=previous_question,
        )
        return source, sql, "openai"

    source = pick_source_for_question(sources, resolved) or sources[0]
    plan = heuristic_plan(source, resolved)
    if not plan["sql"]:
        raise ValueError(
            "No OpenAI API key configured and could not build a heuristic query. "
            "Set it in Settings or OPENAI_API_KEY in .env"
        )
    return source, plan["sql"], ("heuristic" if plan["targeted"] else UNTARGETED)


async def _openai_sql(
    source: DataSource,
    question: str,
    *,
    api_key: str,
    model: str,
    base_url: str,
    previous_question: str | None = None,
) -> str:
    schema_text = build_schema_prompt(source)
    ask = question_prompt_block(question, previous_question)
    content = await _chat_completion(
        system=SYSTEM_PROMPT,
        user=f"{schema_text}\n\n{ask}\n\nSQL:",
        api_key=api_key,
        model=model,
        base_url=base_url,
    )
    return _strip_sql_fences(content)


async def _openai_workspace_sql(
    sources: list[DataSource],
    question: str,
    *,
    api_key: str,
    model: str,
    base_url: str,
    previous_question: str | None = None,
) -> tuple[int, str]:
    catalog = build_workspace_schema_prompt(sources)
    ask = question_prompt_block(question, previous_question)
    content = await _chat_completion(
        system=WORKSPACE_SYSTEM_PROMPT,
        user=f"{catalog}\n\n{ask}",
        api_key=api_key,
        model=model,
        base_url=base_url,
    )
    content = _strip_sql_fences(content)
    source_match = re.search(r"SOURCE_ID\s*:\s*(\d+)", content, flags=re.IGNORECASE)
    sql_match = re.search(r"SQL\s*:\s*(.+)", content, flags=re.IGNORECASE | re.DOTALL)
    resolved = expand_question_with_context(question, previous_question)
    if not source_match or not sql_match:
        picked = pick_source_for_question(sources, resolved) or sources[0]
        return picked.id, content.strip()
    return int(source_match.group(1)), sql_match.group(1).strip()


async def _chat_completion(
    *,
    system: str,
    user: str,
    api_key: str,
    model: str,
    base_url: str,
) -> str:
    payload = {
        "model": model,
        "temperature": 0,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    async with httpx.AsyncClient(timeout=60.0) as client:
        res = await client.post(
            f"{base_url.rstrip('/')}/chat/completions",
            headers=headers,
            json=payload,
        )
        res.raise_for_status()
        data = res.json()
    return data["choices"][0]["message"]["content"].strip()


def _strip_sql_fences(content: str) -> str:
    content = re.sub(r"^```(?:sql)?\s*", "", content, flags=re.IGNORECASE)
    content = re.sub(r"\s*```$", "", content)
    return content.strip()


async def execute_sql(
    source: DataSource, sql: str, *, max_rows: int | None = None
) -> dict[str, Any]:
    safe = validate_readonly_sql(sql)
    safe = ensure_limit(safe, max_rows or settings.max_query_rows)

    if source.source_type == "mysql":
        return await _exec_mysql(source, safe)
    if source.source_type == "file":
        return _exec_file_sqlite(source, safe)
    raise ValueError(f"Unsupported source type for queries: {source.source_type}")


async def _exec_mysql(source: DataSource, sql: str) -> dict[str, Any]:
    import aiomysql

    config = parse_connection_config(source)
    conn = await aiomysql.connect(
        host=config.get("host", "localhost"),
        port=int(config.get("port", 3306)),
        user=config.get("user", "root"),
        password=config.get("password", ""),
        db=config.get("database"),
    )
    try:
        async with conn.cursor() as cur:
            await cur.execute(sql)
            rows_raw = await cur.fetchall()
            columns = [d[0] for d in (cur.description or [])]
            rows = [dict(zip(columns, row, strict=False)) for row in rows_raw]
            # Serialize non-JSON types
            for row in rows:
                for k, v in list(row.items()):
                    if hasattr(v, "isoformat"):
                        row[k] = v.isoformat()
                    elif isinstance(v, (bytes, bytearray)):
                        row[k] = v.decode("utf-8", errors="replace")
    finally:
        conn.close()

    return {"columns": columns, "rows": rows, "sql": sql}


def _exec_file_sqlite(source: DataSource, sql: str) -> dict[str, Any]:
    config = parse_connection_config(source)
    path = Path(config["file_path"])
    if not path.exists():
        raise ValueError("Source file not found on disk")

    table = Path(config.get("original_name", path.name)).stem
    table = re.sub(r"[^A-Za-z0-9_]", "_", table) or "data"
    if table[0].isdigit():
        table = f"t_{table}"

    conn = sqlite3.connect(":memory:")
    try:
        if config.get("format") == "xlsx":
            _load_xlsx_into_sqlite(conn, path, table)
        else:
            _load_csv_into_sqlite(conn, path, table)

        # Rewrite common original table names to sanitized name
        rewritten = sql
        original = Path(config.get("original_name", path.name)).stem
        if original and original != table:
            rewritten = re.sub(
                rf"(?i)`?{re.escape(original)}`?",
                f"`{table}`",
                rewritten,
            )

        cur = conn.execute(rewritten)
        rows_raw = cur.fetchall()
        columns = [d[0] for d in (cur.description or [])]
        rows = [dict(zip(columns, row, strict=False)) for row in rows_raw]
    finally:
        conn.close()

    return {"columns": columns, "rows": rows, "sql": sql}


def _load_csv_into_sqlite(conn: sqlite3.Connection, path: Path, table: str) -> None:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            raise ValueError("CSV has no header row")
        cols = [_safe_col(h, i) for i, h in enumerate(header)]
        raw_rows: list[list[str]] = []
        for row in reader:
            padded = (row + [""] * len(cols))[: len(cols)]
            raw_rows.append(padded)
        _insert_typed_rows(conn, table, cols, raw_rows)


def _load_xlsx_into_sqlite(conn: sqlite3.Connection, path: Path, table: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        wb.close()
        raise ValueError("Excel sheet has no header row")
    cols = [_safe_col(str(h) if h is not None else "", i) for i, h in enumerate(header)]
    raw_rows: list[list[str]] = []
    for row in rows_iter:
        padded = []
        for i in range(len(cols)):
            val = row[i] if i < len(row) else None
            padded.append("" if val is None else str(val))
        raw_rows.append(padded)
    wb.close()
    _insert_typed_rows(conn, table, cols, raw_rows)


def _insert_typed_rows(
    conn: sqlite3.Connection,
    table: str,
    cols: list[str],
    raw_rows: list[list[str]],
) -> None:
    """Infer REAL vs TEXT per column so ORDER BY works on numeric CSV fields."""
    types = [_infer_sqlite_type(raw_rows, i) for i in range(len(cols))]
    col_defs = ", ".join(f"`{c}` {t}" for c, t in zip(cols, types, strict=False))
    conn.execute(f"CREATE TABLE `{table}` ({col_defs})")
    placeholders = ", ".join("?" for _ in cols)
    typed_rows: list[list[Any]] = []
    for row in raw_rows:
        typed: list[Any] = []
        for i, val in enumerate(row):
            if types[i] == "REAL":
                typed.append(None if val == "" else float(val.replace(",", "")))
            else:
                typed.append(val)
        typed_rows.append(typed)
    conn.executemany(f"INSERT INTO `{table}` VALUES ({placeholders})", typed_rows)
    conn.commit()


def _infer_sqlite_type(rows: list[list[str]], col_index: int) -> str:
    seen = False
    for row in rows:
        val = row[col_index].strip() if col_index < len(row) else ""
        if val == "":
            continue
        seen = True
        if not re.fullmatch(r"-?\d+(?:\.\d+)?", val.replace(",", "")):
            return "TEXT"
    return "REAL" if seen else "TEXT"


def _safe_col(name: str, index: int) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name.strip()) or f"column_{index + 1}"
    if cleaned[0].isdigit():
        cleaned = f"c_{cleaned}"
    return cleaned


def pack_result(result: dict[str, Any]) -> str:
    return json.dumps(result, default=str)


async def generate_narrative(
    question: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    *,
    api_key: str | None = None,
    model: str | None = None,
    base_url: str | None = None,
    source_name: str | None = None,
    currency: str | None = None,
) -> str | None:
    """Ask the model to summarise the returned rows. None when unavailable.

    The prompt carries the actual rows, so the summary stays grounded in the
    query result rather than the model's own recollection.
    """
    key = api_key if api_key is not None else settings.openai_api_key
    if not key or not rows:
        return None

    mdl = model or settings.openai_model
    url = base_url or settings.openai_base_url
    prompt = build_narrative_prompt(
        question, columns, rows, source_name=source_name, currency=currency
    )

    payload = {
        "model": mdl,
        "temperature": 0.2,
        "max_tokens": 220,
        "messages": [
            {"role": "system", "content": NARRATIVE_SYSTEM},
            {"role": "user", "content": prompt},
        ],
    }
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            res = await client.post(
                f"{url.rstrip('/')}/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )
            res.raise_for_status()
            data = res.json()
        content = data["choices"][0]["message"]["content"]
    except Exception:
        # A narrative is a nicety; the deterministic summary already covers it.
        return None

    cleaned = sanitize_narrative(content)
    return cleaned or None


async def generate_diagnostic_answer(
    question: str,
    diagnosis: dict[str, Any],
    *,
    advisory: bool,
    candidates: list[dict[str, str]] | None = None,
    api_key: str | None = None,
    model: str | None = None,
    base_url: str | None = None,
    source_name: str | None = None,
    currency: str | None = None,
) -> tuple[str | None, list[dict[str, str]]]:
    """Write up a measured diagnosis, and for advisory questions, the actions.

    The model is given the computed evidence, never the raw rows, so it has
    nothing to miscalculate. Returns (None, []) whenever it is unavailable or
    replies with anything unusable — the caller then uses the deterministic
    write-up, which says the same thing from the same numbers.
    """
    key = api_key if api_key is not None else settings.openai_api_key
    if not key:
        return None, []

    mdl = model or settings.openai_model
    url = base_url or settings.openai_base_url
    prompt = build_evidence_prompt(
        question,
        diagnosis,
        source_name=source_name,
        currency=currency,
        candidates=candidates if advisory else None,
    )

    payload = {
        "model": mdl,
        "temperature": 0.2,
        "max_tokens": 700 if advisory else 300,
        "messages": [
            {"role": "system", "content": ADVISORY_SYSTEM if advisory else DIAGNOSTIC_SYSTEM},
            {"role": "user", "content": prompt},
        ],
    }
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            res = await client.post(
                f"{url.rstrip('/')}/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )
            res.raise_for_status()
            data = res.json()
        content = data["choices"][0]["message"]["content"]
    except Exception:
        return None, []

    if advisory:
        answer, recommendations = parse_advisory_json(content)
        return answer, recommendations

    cleaned = sanitize_narrative(content)
    return (cleaned or None), []


async def generate_partial_answer(
    question: str,
    context: dict[str, Any],
    *,
    api_key: str | None = None,
    model: str | None = None,
    base_url: str | None = None,
    source_name: str | None = None,
    currency: str | None = None,
) -> str | None:
    """Answer a "why" question the data cannot fully support.

    The model is handed the comparison that does exist and an explicit list of
    what is missing, so it can say what is answerable instead of the question
    dead-ending. None when unavailable; the caller writes it up deterministically.
    """
    key = api_key if api_key is not None else settings.openai_api_key
    if not key:
        return None

    payload = {
        "model": model or settings.openai_model,
        "temperature": 0.2,
        "max_tokens": 320,
        "messages": [
            {"role": "system", "content": PARTIAL_SYSTEM},
            {
                "role": "user",
                "content": build_partial_prompt(
                    question, context, source_name=source_name, currency=currency
                ),
            },
        ],
    }
    url = base_url or settings.openai_base_url
    try:
        async with httpx.AsyncClient(timeout=45.0) as client:
            res = await client.post(
                f"{url.rstrip('/')}/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "Content-Type": "application/json",
                },
                json=payload,
            )
            res.raise_for_status()
            data = res.json()
        content = data["choices"][0]["message"]["content"]
    except Exception:
        return None

    return sanitize_narrative(content) or None

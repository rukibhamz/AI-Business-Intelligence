from __future__ import annotations

import csv
import json
import uuid
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from app.services.schema_types import ColumnSchema, SourceSchema, TableSchema


def infer_type(value: str) -> str:
    if value == "":
        return "string"
    lowered = value.lower()
    if lowered in ("true", "false"):
        return "boolean"
    try:
        int(value)
        return "integer"
    except ValueError:
        pass
    try:
        float(value)
        return "number"
    except ValueError:
        pass
    return "string"


def merge_column_types(existing: str, new: str) -> str:
    if existing == new:
        return existing
    if {existing, new} <= {"integer", "number"}:
        return "number"
    return "string"


class BaseConnector(ABC):
    @abstractmethod
    async def introspect(self) -> SourceSchema:
        raise NotImplementedError

    @abstractmethod
    async def preview(
        self, *, table: str | None = None, limit: int = 100, offset: int = 0
    ) -> dict[str, Any]:
        raise NotImplementedError


class FileConnector(BaseConnector):
    def __init__(self, file_path: Path, original_name: str, file_format: str) -> None:
        self.file_path = file_path
        self.original_name = original_name
        self.file_format = file_format

    async def introspect(self) -> SourceSchema:
        if self.file_format == "csv":
            columns = self._introspect_csv()
        elif self.file_format == "xlsx":
            columns = self._introspect_xlsx()
        else:
            raise ValueError(f"Unsupported file format: {self.file_format}")

        table_name = Path(self.original_name).stem
        return {"tables": [{"name": table_name, "columns": columns}]}

    async def preview(
        self, *, table: str | None = None, limit: int = 100, offset: int = 0
    ) -> dict[str, Any]:
        if self.file_format == "csv":
            rows, columns, total = self._preview_csv(limit, offset)
        elif self.file_format == "xlsx":
            rows, columns, total = self._preview_xlsx(limit, offset)
        else:
            raise ValueError(f"Unsupported file format: {self.file_format}")

        table_name = Path(self.original_name).stem
        return {
            "table": table or table_name,
            "columns": columns,
            "rows": rows,
            "limit": limit,
            "offset": offset,
            "total": total,
        }

    def _introspect_csv(self) -> list[ColumnSchema]:
        with self.file_path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if not header:
                return []

            types: list[str | None] = [None] * len(header)
            for row in reader:
                for i, cell in enumerate(row[: len(header)]):
                    inferred = infer_type(cell)
                    if types[i] is None:
                        types[i] = inferred
                    else:
                        types[i] = merge_column_types(types[i], inferred)
            return [
                {"name": col.strip() or f"column_{idx + 1}", "type": types[idx] or "string"}
                for idx, col in enumerate(header)
            ]

    def _preview_csv(self, limit: int, offset: int) -> tuple[list[dict[str, Any]], list[str], int]:
        with self.file_path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None) or []
            columns = [col.strip() or f"column_{i + 1}" for i, col in enumerate(header)]
            data_rows: list[list[str]] = list(reader)

        total = len(data_rows)
        page = data_rows[offset : offset + limit]
        rows = [dict(zip(columns, row + [""] * (len(columns) - len(row)))) for row in page]
        return rows, columns, total

    def _introspect_xlsx(self) -> list[ColumnSchema]:
        from openpyxl import load_workbook

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return []

        columns = [str(h).strip() if h is not None else f"column_{i + 1}" for i, h in enumerate(header)]
        types: list[str | None] = [None] * len(columns)

        for row in rows_iter:
            for i, cell in enumerate(row[: len(columns)]):
                value = "" if cell is None else str(cell)
                inferred = infer_type(value)
                if types[i] is None:
                    types[i] = inferred
                else:
                    types[i] = merge_column_types(types[i], inferred)

        wb.close()
        return [{"name": columns[i], "type": types[i] or "string"} for i in range(len(columns))]

    def _preview_xlsx(self, limit: int, offset: int) -> tuple[list[dict[str, Any]], list[str], int]:
        from openpyxl import load_workbook

        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return [], [], 0

        header = all_rows[0]
        columns = [str(h).strip() if h is not None else f"column_{i + 1}" for i, h in enumerate(header)]
        data_rows = all_rows[1:]
        total = len(data_rows)
        page = data_rows[offset : offset + limit]
        rows = [
            {
                columns[i]: ("" if i >= len(row) or row[i] is None else row[i])
                for i in range(len(columns))
            }
            for row in page
        ]
        return rows, columns, total


class MySQLConnector(BaseConnector):
    def __init__(self, config: dict[str, Any]) -> None:
        self.config = config

    def _connect_kwargs(self) -> dict[str, Any]:
        return {
            "host": self.config.get("host", "localhost"),
            "port": int(self.config.get("port", 3306)),
            "user": self.config.get("user", "root"),
            "password": self.config.get("password", ""),
            "db": self.config.get("database"),
        }

    async def test_connection(self) -> None:
        import aiomysql

        conn = await aiomysql.connect(**self._connect_kwargs())
        conn.close()

    async def introspect(self) -> SourceSchema:
        import aiomysql

        conn = await aiomysql.connect(**self._connect_kwargs())
        try:
            async with conn.cursor() as cur:
                await cur.execute("SHOW TABLES")
                tables_raw = await cur.fetchall()
                tables: list[TableSchema] = []
                for (table_name,) in tables_raw:
                    await cur.execute(f"DESCRIBE `{table_name}`")
                    cols = await cur.fetchall()
                    columns = [{"name": row[0], "type": self._map_mysql_type(row[1])} for row in cols]
                    tables.append({"name": table_name, "columns": columns})
        finally:
            conn.close()

        return {"tables": tables}

    async def preview(
        self, *, table: str | None = None, limit: int = 100, offset: int = 0
    ) -> dict[str, Any]:
        import aiomysql

        if not table:
            schema = await self.introspect()
            if not schema["tables"]:
                return {"table": None, "columns": [], "rows": [], "limit": limit, "offset": offset, "total": 0}
            table = schema["tables"][0]["name"]

        conn = await aiomysql.connect(**self._connect_kwargs())
        try:
            async with conn.cursor() as cur:
                await cur.execute(f"SELECT COUNT(*) FROM `{table}`")
                total = (await cur.fetchone())[0]
                await cur.execute(f"SELECT * FROM `{table}` LIMIT %s OFFSET %s", (limit, offset))
                rows_raw = await cur.fetchall()
                columns = [desc[0] for desc in cur.description or []]
                rows = [dict(zip(columns, row)) for row in rows_raw]
        finally:
            conn.close()

        return {
            "table": table,
            "columns": columns,
            "rows": rows,
            "limit": limit,
            "offset": offset,
            "total": total,
        }

    @staticmethod
    def _map_mysql_type(mysql_type: str) -> str:
        t = mysql_type.lower()
        if any(k in t for k in ("int", "decimal", "float", "double", "numeric")):
            return "number" if "int" not in t else "integer"
        if any(k in t for k in ("date", "time", "year")):
            return "datetime"
        if "bool" in t or "bit" in t:
            return "boolean"
        return "string"


def detect_file_format(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext in (".xlsx", ".xls"):
        return "xlsx"
    raise ValueError("Supported file types: .csv, .xlsx")


def save_upload(upload_dir: Path, filename: str, content: bytes) -> tuple[Path, str]:
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}_{Path(filename).name}"
    dest = upload_dir / safe_name
    dest.write_bytes(content)
    return dest, detect_file_format(filename)


def build_file_connector(config: dict[str, Any]) -> FileConnector:
    return FileConnector(
        file_path=Path(config["file_path"]),
        original_name=config.get("original_name", Path(config["file_path"]).name),
        file_format=config.get("format", "csv"),
    )


def build_mysql_connector(config: dict[str, Any]) -> MySQLConnector:
    return MySQLConnector(config)
